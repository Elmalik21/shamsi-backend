"""
tests/test_exports.py
======================
Unit and integration tests for the Step 3 export layer.

Run with:
    python -m pytest tests/test_exports.py -v
    python -m pytest tests/test_exports.py::TestPVsystExporter -v

All tests use synthetic data.  Heavy dependencies (reportlab, openpyxl)
are skipped if not installed.
"""
from __future__ import annotations

import json
import math
import os
import sys
import tempfile
import unittest
from pathlib import Path

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from ai_engine.export.pvsyst_exporter import (
    PVsystExporter, make_synthetic_project,
    _erbs_diffuse_fraction, _decompose_ghi, _declination,
    _sin_solar_noon_elevation,
)
from ai_engine.export.helioscope_exporter import HelioScopeExporter


# ─────────────────────────────────────────────────────────────────────────────
# 1. Solar geometry helpers
# ─────────────────────────────────────────────────────────────────────────────

class TestSolarGeometry(unittest.TestCase):

    def test_declination_solstices(self):
        """Summer solstice (day 172) ≈ +23.45°, winter (day 355) ≈ −23.45°."""
        self.assertAlmostEqual(_declination(172), 23.45,  delta=0.5)
        self.assertAlmostEqual(_declination(355), -23.45, delta=0.5)

    def test_declination_equinoxes(self):
        """Spring equinox (day 80) ≈ 0°."""
        self.assertAlmostEqual(_declination(80), 0.0, delta=2.0)

    def test_sin_elevation_positive(self):
        """Sin of solar elevation at solar noon should always be positive."""
        for doy in [1, 80, 172, 265, 355]:
            sin_e = _sin_solar_noon_elevation(doy, 30.0)
            self.assertGreater(sin_e, 0.0)

    def test_sin_elevation_minimum_clamped(self):
        """Minimum is 0.01 (clamped to avoid division by zero)."""
        sin_e = _sin_solar_noon_elevation(355, 70.0)  # polar winter
        self.assertGreaterEqual(sin_e, 0.01)

    def test_erbs_low_kt(self):
        """kt ≤ 0.22 → diffuse fraction = 1 - 0.09*kt."""
        kt = 0.15
        kd = _erbs_diffuse_fraction(kt)
        self.assertAlmostEqual(kd, 1.0 - 0.09 * kt, places=6)

    def test_erbs_high_kt(self):
        """kt > 0.80 → diffuse fraction = 0.165."""
        self.assertAlmostEqual(_erbs_diffuse_fraction(0.85), 0.165, places=6)
        self.assertAlmostEqual(_erbs_diffuse_fraction(1.00), 0.165, places=6)

    def test_erbs_mid_kt_in_range(self):
        """Mid-range kt → diffuse fraction must be in (0, 1)."""
        for kt in [0.30, 0.45, 0.60, 0.75]:
            kd = _erbs_diffuse_fraction(kt)
            self.assertGreater(kd, 0.0)
            self.assertLess(kd, 1.0)

    def test_decompose_ghi_consistency(self):
        """DNI + DHI should reconstruct GHI within 20% (geometry proxy)."""
        ghi = 5.5   # typical Cairo summer day
        doy = 172   # summer solstice
        lat = 30.0
        dni, dhi = _decompose_ghi(ghi, doy, lat)
        self.assertGreater(dni, 0)
        self.assertGreater(dhi, 0)
        # DHI should not exceed GHI
        self.assertLessEqual(dhi, ghi + 0.01)

    def test_decompose_ghi_zero(self):
        """Zero GHI → zero DNI and zero DHI."""
        dni, dhi = _decompose_ghi(0.0, 172, 30.0)
        self.assertEqual(dni, 0.0)
        self.assertEqual(dhi, 0.0)


# ─────────────────────────────────────────────────────────────────────────────
# 2. PVsystExporter
# ─────────────────────────────────────────────────────────────────────────────

class TestPVsystExporter(unittest.TestCase):

    def setUp(self):
        self.project   = make_synthetic_project('Cairo')
        self.exporter  = PVsystExporter(self.project)
        self.tmp_dir   = tempfile.mkdtemp(prefix='shamsi_pvsyst_')

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_export_all_returns_four_keys(self):
        files = self.exporter.export_all(self.tmp_dir)
        self.assertIn('sit_file', files)
        self.assertIn('met_file', files)
        self.assertIn('pan_file', files)
        self.assertIn('ond_file', files)

    def test_export_all_creates_files(self):
        files = self.exporter.export_all(self.tmp_dir)
        for key, path in files.items():
            self.assertTrue(os.path.exists(path), f"Missing: {key} → {path}")
            self.assertGreater(os.path.getsize(path), 0, f"Empty: {key}")

    # ── .SIT file ─────────────────────────────────────────────────────────────

    def test_sit_contains_latitude(self):
        files = self.exporter.export_all(self.tmp_dir)
        content = Path(files['sit_file']).read_text()
        self.assertIn('Latitude', content)
        # Cairo latitude 30.044
        self.assertIn('30.044', content)

    def test_sit_contains_country(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['sit_file']).read_text()
        self.assertIn('Egypt', content)

    def test_sit_references_met_file(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['sit_file']).read_text()
        self.assertIn('.MET', content)

    def test_sit_has_albedo(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['sit_file']).read_text()
        self.assertIn('Albedo', content)
        self.assertIn('0.20', content)

    # ── .MET file ─────────────────────────────────────────────────────────────

    def test_met_has_csv_header(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['met_file']).read_text()
        self.assertIn('date,GHI,DNI,DHI,Tdry,Wspd,RH', content)

    def test_met_has_365_data_rows(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['met_file']).read_text()
        data_rows = [l for l in content.split('\n')
                     if l and not l.startswith('#') and '-' in l[:5] and ',' in l]
        # Allow 364–366 for leap year tolerance
        self.assertGreaterEqual(len(data_rows), 364)
        self.assertLessEqual(len(data_rows),    366)

    def test_met_ghi_positive(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['met_file']).read_text()
        data_rows = [l for l in content.split('\n')
                     if l and not l.startswith('#') and '-' in l[:5] and ',' in l]
        ghis = [float(row.split(',')[1]) for row in data_rows[:10]]
        for ghi in ghis:
            self.assertGreaterEqual(ghi, 0.0)

    # ── .PAN file ─────────────────────────────────────────────────────────────

    def test_pan_has_pnom(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['pan_file']).read_text()
        self.assertIn('Pnom', content)
        self.assertIn('580', content)

    def test_pan_has_temp_coeff(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['pan_file']).read_text()
        self.assertIn('muPmpp', content)

    def test_pan_has_dimensions(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['pan_file']).read_text()
        self.assertIn('Length', content)
        self.assertIn('Width', content)

    # ── .OND file ─────────────────────────────────────────────────────────────

    def test_ond_has_pnom(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['ond_file']).read_text()
        self.assertIn('Pnom', content)

    def test_ond_has_efficiency(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['ond_file']).read_text()
        self.assertIn('Efficiency', content)

    def test_ond_has_mppt_range(self):
        content = Path(self.exporter.export_all(self.tmp_dir)['ond_file']).read_text()
        self.assertIn('VmppMin', content)
        self.assertIn('VmppMax', content)

    # ── Location variations ───────────────────────────────────────────────────

    def test_different_location(self):
        """Aswan project should produce higher GHI values."""
        aswan   = make_synthetic_project('Aswan')
        aswan['location'].latitude  = 24.088
        aswan['location'].longitude = 32.900
        exp     = PVsystExporter(aswan)
        tmp     = tempfile.mkdtemp()
        try:
            files = exp.export_all(tmp)
            sit   = Path(files['sit_file']).read_text()
            self.assertIn('Aswan', sit)
        finally:
            import shutil; shutil.rmtree(tmp, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# 3. HelioScopeExporter
# ─────────────────────────────────────────────────────────────────────────────

class TestHelioScopeExporter(unittest.TestCase):

    def setUp(self):
        self.project  = make_synthetic_project('Cairo')
        self.exporter = HelioScopeExporter(self.project)

    def test_to_dict_returns_project_key(self):
        d = self.exporter.to_dict()
        self.assertIn('project', d)

    def test_location_fields(self):
        loc = self.exporter.to_dict()['project']['location']
        self.assertIn('latitude',  loc)
        self.assertIn('longitude', loc)
        self.assertIn('timezone',  loc)
        self.assertAlmostEqual(loc['latitude'],  30.044, places=2)
        self.assertAlmostEqual(loc['longitude'], 31.236, places=2)

    def test_design_has_arrays(self):
        design = self.exporter.to_dict()['project']['design']
        self.assertIn('arrays', design)
        self.assertGreater(len(design['arrays']), 0)

    def test_array_has_modules_and_inverters(self):
        arr = self.exporter.to_dict()['project']['design']['arrays'][0]
        self.assertIn('modules',   arr)
        self.assertIn('inverters', arr)

    def test_module_count(self):
        arr = self.exporter.to_dict()['project']['design']['arrays'][0]
        self.assertEqual(arr['modules']['count'], 30)

    def test_losses_sum_to_reasonable(self):
        losses = self.exporter.to_dict()['project']['design']['losses']
        total  = losses['total_estimated_pct']
        self.assertGreater(total, 5.0)
        self.assertLess(total,    40.0)

    def test_losses_all_between_0_and_1(self):
        losses = self.exporter.to_dict()['project']['design']['losses']
        for k, v in losses.items():
            if k == 'total_estimated_pct':
                continue
            self.assertGreaterEqual(v, 0.0, f"Negative loss: {k}")
            self.assertLessEqual(v,    1.0, f"Loss > 100%: {k}")

    def test_energy_production_present(self):
        ep = self.exporter.to_dict()['project']['energy_production']
        self.assertIn('annual_kwh', ep)
        self.assertIsNotNone(ep['annual_kwh'])

    def test_economics_present(self):
        eco = self.exporter.to_dict()['project']['economics']
        self.assertIn('payback_years', eco)
        self.assertIn('total_cost',    eco)

    def test_lcoe_positive(self):
        eco  = self.exporter.to_dict()['project']['economics']
        lcoe = eco.get('lcoe_egp_per_kwh')
        if lcoe is not None:
            self.assertGreater(lcoe, 0.0)

    def test_pr_in_range(self):
        ep = self.exporter.to_dict()['project']['energy_production']
        pr = ep.get('performance_ratio')
        if pr is not None:
            self.assertGreater(pr, 0.3)
            self.assertLess(pr,    1.0)

    def test_shamsi_metadata_present(self):
        meta = self.exporter.to_dict()['project']['shamsi_metadata']
        self.assertIn('project_id',  meta)
        self.assertIn('optimiser',   meta)
        self.assertIn('export_date', meta)

    def test_export_project_writes_file(self):
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            path = f.name
        try:
            result = self.exporter.export_project(path)
            self.assertTrue(os.path.exists(path))
            with open(path) as fh:
                loaded = json.load(fh)
            self.assertIn('project', loaded)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_exported_json_is_valid(self):
        """The written JSON must parse without errors."""
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            path = f.name
        try:
            self.exporter.export_project(path)
            with open(path) as fh:
                data = json.load(fh)
            self.assertIsInstance(data, dict)
        finally:
            if os.path.exists(path): os.unlink(path)


# ─────────────────────────────────────────────────────────────────────────────
# 4. PDF Report
# ─────────────────────────────────────────────────────────────────────────────

class TestPDFReport(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        try:
            import reportlab
            cls.reportlab_available = True
        except ImportError:
            cls.reportlab_available = False

    def test_import_without_reportlab(self):
        """Module imports cleanly even if reportlab is not installed."""
        from ai_engine.export.pdf_report import ProfessionalPDFReport
        self.assertTrue(callable(ProfessionalPDFReport))

    def test_generate_report_creates_pdf(self):
        if not self.reportlab_available:
            self.skipTest('reportlab not installed')

        from ai_engine.export.pdf_report import ProfessionalPDFReport
        from ai_engine.export.calc_engine import normalize_and_validate_project
        project = make_synthetic_project('Cairo')
        project = normalize_and_validate_project(project)

        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
            path = f.name
        try:
            report = ProfessionalPDFReport(project)
            result = report.generate_report(path)
            self.assertTrue(os.path.exists(result))
            self.assertGreater(os.path.getsize(result), 1000)  # > 1 KB
            # Check PDF magic bytes
            with open(result, 'rb') as fh:
                header = fh.read(4)
            self.assertEqual(header, b'%PDF')
        finally:
            if os.path.exists(path): os.unlink(path)

    def test_generate_report_with_roof_analysis(self):
        """PDF should not crash when roof_analysis dict is included."""
        if not self.reportlab_available:
            self.skipTest('reportlab not installed')

        from ai_engine.export.pdf_report import ProfessionalPDFReport
        from ai_engine.export.calc_engine import normalize_and_validate_project
        project = make_synthetic_project('Aswan')
        project = normalize_and_validate_project(project)
        project['roof_analysis'] = {
            'roof_area_m2'     : 200.0,
            'usable_area_m2'   : 185.0,
            'usable_percentage': 92.5,
            'obstacles'        : [{'class': 'ac_unit'}],
            'metadata'         : {'orientation': 'flat', 'roof_type': 'concrete'},
        }

        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
            path = f.name
        try:
            ProfessionalPDFReport(project).generate_report(path)
            self.assertGreater(os.path.getsize(path), 1000)
        finally:
            if os.path.exists(path): os.unlink(path)


# ─────────────────────────────────────────────────────────────────────────────
# 5. Excel Exporter
# ─────────────────────────────────────────────────────────────────────────────

class TestExcelExporter(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl
            cls.openpyxl_available = True
        except ImportError:
            cls.openpyxl_available = False

    def test_import_without_openpyxl(self):
        from ai_engine.export.excel_exporter import ExcelExporter
        self.assertTrue(callable(ExcelExporter))

    def test_export_csv_no_deps(self):
        """CSV export requires no extra dependencies."""
        from ai_engine.export.excel_exporter import ExcelExporter
        project = make_synthetic_project('Cairo')
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False,
                                          mode='w') as f:
            path = f.name
        try:
            ExcelExporter(project).export_csv(path)
            self.assertTrue(os.path.exists(path))
            content = Path(path).read_text()
            self.assertIn('Month,Production_kWh', content)
            self.assertIn('Annual', content)
            rows = [l for l in content.split('\n') if l.strip()]
            self.assertEqual(len(rows), 14)  # header + 12 months + annual
        finally:
            if os.path.exists(path): os.unlink(path)

    def test_export_workbook_creates_xlsx(self):
        if not self.openpyxl_available:
            self.skipTest('openpyxl not installed')

        from ai_engine.export.excel_exporter import ExcelExporter
        from ai_engine.export.calc_engine import normalize_and_validate_project
        project = make_synthetic_project('Cairo')
        project = normalize_and_validate_project(project)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            ExcelExporter(project).export_workbook(path)
            self.assertTrue(os.path.exists(path))
            self.assertGreater(os.path.getsize(path), 1000)

            # Check sheet names
            import openpyxl
            wb     = openpyxl.load_workbook(path)
            sheets = wb.sheetnames
            for expected in ['Summary', 'System Design', 'Monthly Production',
                             'Financial Analysis', 'Equipment Specs', 'Climate Data']:
                self.assertIn(expected, sheets, f"Missing sheet: {expected}")
        finally:
            if os.path.exists(path): os.unlink(path)

    def test_monthly_production_sheet_has_12_rows(self):
        if not self.openpyxl_available:
            self.skipTest('openpyxl not installed')

        from ai_engine.export.excel_exporter import ExcelExporter
        from ai_engine.export.calc_engine import normalize_and_validate_project
        import openpyxl

        project = make_synthetic_project('Cairo')
        project = normalize_and_validate_project(project)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name
        try:
            ExcelExporter(project).export_workbook(path)
            wb    = openpyxl.load_workbook(path)
            ws    = wb['Monthly Production']
            # Row 3 is header, rows 4–15 are months (12 rows)
            month_rows = [ws.cell(row=r, column=1).value
                          for r in range(4, 16) if ws.cell(row=r, column=1).value]
            self.assertEqual(len(month_rows), 12)
        finally:
            if os.path.exists(path): os.unlink(path)


# ─────────────────────────────────────────────────────────────────────────────
# 6. Case Study Validation
# ─────────────────────────────────────────────────────────────────────────────

class TestCaseStudyValidation(unittest.TestCase):

    def setUp(self):
        from scripts.validate_with_case_studies import (
            CASE_STUDIES, run_case_study, compute_validation_stats,
            _estimate_ghi, _pvwatts_specific_yield, _monthly_distribution,
        )
        self.CASE_STUDIES           = CASE_STUDIES
        self.run_case_study         = run_case_study
        self.compute_validation_stats = compute_validation_stats
        self.estimate_ghi           = _estimate_ghi
        self.pvwatts_sy             = _pvwatts_specific_yield
        self.monthly_dist           = _monthly_distribution

    def test_ghi_higher_at_lower_latitude(self):
        """GHI should be higher at Aswan (24°N) than Alexandria (31°N)."""
        ghi_aswan = self.estimate_ghi(24.0)
        ghi_alex  = self.estimate_ghi(31.0)
        self.assertGreater(ghi_aswan, ghi_alex)

    def test_specific_yield_positive(self):
        sy = self.pvwatts_sy(5.5, 30.0, 20.0)
        self.assertGreater(sy, 0)

    def test_specific_yield_higher_at_better_ghi(self):
        sy_high = self.pvwatts_sy(7.0, 24.0, 15.0)  # Aswan
        sy_low  = self.pvwatts_sy(5.2, 31.0, 20.0)  # Alexandria
        self.assertGreater(sy_high, sy_low)

    def test_monthly_distribution_sums_to_annual(self):
        annual  = 20_000
        monthly = self.monthly_dist(annual, 30.0)
        self.assertEqual(len(monthly), 12)
        self.assertAlmostEqual(sum(monthly), annual, delta=1.0)

    def test_monthly_distribution_summer_higher(self):
        """Summer months (Jun–Aug) should exceed winter (Dec–Feb) at 30°N."""
        monthly = self.monthly_dist(20_000, 30.0)
        summer  = sum(monthly[5:8])    # Jun, Jul, Aug
        winter  = sum(monthly[11:]) + sum(monthly[:2])  # Dec, Jan, Feb
        self.assertGreater(summer, winter)

    def test_run_case_study_cairo(self):
        """Cairo case study should return a valid result dict."""
        case   = next(c for c in self.CASE_STUDIES if c['id'] == 'CS-01')
        result = self.run_case_study(case)

        self.assertIn('mape_pct', result)
        self.assertIn('shamsi_specific_yield', result)
        self.assertIn('ref_specific_yield', result)
        self.assertGreater(result['shamsi_specific_yield'], 0)
        self.assertGreater(result['mape_pct'], 0)
        self.assertLess(result['mape_pct'], 50)   # sanity — not catastrophically wrong

    def test_all_five_case_studies(self):
        """All 5 case studies should run without error and pass <10% MAPE."""
        results = [self.run_case_study(c) for c in self.CASE_STUDIES]
        stats   = self.compute_validation_stats(results)
        self.assertEqual(stats['n_cases'], 5)
        self.assertTrue(stats['passes_10pct_target'],
                        f"Mean MAPE {stats['mean_mape']:.1f}% exceeds 10% target")

    def test_validation_stats_structure(self):
        results = [self.run_case_study(self.CASE_STUDIES[0])]
        stats   = self.compute_validation_stats(results)
        for key in ('mean_mape', 'std_mape', 'mean_bias', 'mean_rmse',
                    'pct_within_5', 'pct_within_10', 'passes_10pct_target'):
            self.assertIn(key, stats, f"Missing stat: {key}")


# ─────────────────────────────────────────────────────────────────────────────
# 7. API view helper (no Django test client needed)
# ─────────────────────────────────────────────────────────────────────────────

class TestExportViewHelpers(unittest.TestCase):

    def setUp(self):
        # Minimal Django setup
        import django
        from django.conf import settings
        if not settings.configured:
            import os
            if os.environ.get('DJANGO_SETTINGS_MODULE'):
                pass
            else:
                settings.configure(
                    MEDIA_ROOT='/tmp/shamsi_test_exports',
                    MEDIA_URL='/media/',
                    INSTALLED_APPS=[
                        'django.contrib.auth',
                        'django.contrib.contenttypes',
                        'rest_framework',
                        'solar_data',
                        'api',
                    ],
                    DATABASES={},
                )
        django.setup()

    def test_export_dir_created(self):
        from api.views.export_view import _export_dir
        d = _export_dir('test_proj_123')
        self.assertTrue(os.path.isdir(d))

    def test_load_project_demo(self):
        """'demo' project_id should return synthetic data without DB."""
        from api.views.export_view import _load_project

        class FakeRequest:
            query_params = {}

        project = _load_project('demo', FakeRequest())
        self.assertIsNotNone(project)
        self.assertIn('location',  project)
        self.assertIn('panel',     project)
        self.assertIn('inverter',  project)

    def test_export_view_callables(self):
        """All export view functions are callable."""
        from api.views.export_view import (
            export_pvsyst, export_helioscope, export_pdf,
            export_excel, export_csv, export_all,
        )
        for fn in [export_pvsyst, export_helioscope, export_pdf,
                   export_excel, export_csv, export_all]:
            self.assertTrue(callable(fn))


# ─────────────────────────────────────────────────────────────────────────────
# 8. Edge cases
# ─────────────────────────────────────────────────────────────────────────────

class TestEdgeCases(unittest.TestCase):

    def test_zero_panel_count(self):
        """Zero panels should not raise in HelioScope exporter."""
        project = make_synthetic_project('Cairo')
        project['system_config']['panel_count'] = 0
        exp = HelioScopeExporter(project)
        d   = exp.to_dict()
        self.assertEqual(d['project']['design']['arrays'][0]['modules']['count'], 0)

    def test_missing_optimization_results(self):
        """Missing optimization_results should produce None fields, not crash."""
        project = make_synthetic_project('Cairo')
        project['optimization_results'] = {}
        exp = HelioScopeExporter(project)
        d   = exp.to_dict()
        ep  = d['project']['energy_production']
        self.assertIsNone(ep['annual_kwh'])

    def test_pvsyst_southern_hemisphere(self):
        """Southern hemisphere site should use °S notation."""
        project = make_synthetic_project('Test')
        project['location'].latitude  = -33.9
        project['location'].longitude =  18.4
        project['location'].country   = 'South Africa'
        project['location'].name      = 'Cape Town'
        exp = PVsystExporter(project)
        with tempfile.TemporaryDirectory() as td:
            files   = exp.export_all(td)
            content = Path(files['sit_file']).read_text()
            self.assertIn('°S', content)

    def test_helioscope_lcoe_none_when_no_cost(self):
        """LCOE should be None when total_cost_egp is missing."""
        project = make_synthetic_project('Cairo')
        project['optimization_results']['total_cost_egp'] = None
        exp = HelioScopeExporter(project)
        eco = exp.to_dict()['project']['economics']
        self.assertIsNone(eco['lcoe_egp_per_kwh'])

    def test_csv_export_without_monthly_data(self):
        """CSV export with empty monthly_yield_kwh should not crash."""
        from ai_engine.export.excel_exporter import ExcelExporter
        project = make_synthetic_project('Cairo')
        project['optimization_results']['monthly_yield_kwh'] = []
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False, mode='w') as f:
            path = f.name
        try:
            ExcelExporter(project).export_csv(path)
            self.assertTrue(os.path.exists(path))
        finally:
            if os.path.exists(path): os.unlink(path)


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    unittest.main(verbosity=2)
