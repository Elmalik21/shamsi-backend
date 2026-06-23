"""
ai_engine/export/excel_exporter.py
=====================================
Export Shamsi Smart design results to a formatted Excel workbook.

Sheets produced
---------------
  1. Summary              — key KPIs with conditional formatting
  2. System Design        — full equipment specifications
  3. Monthly Production   — 12-month table + embedded bar chart
  4. Financial Analysis   — 25-year cashflow table + summary
  5. Equipment Specs      — panel and inverter data in detail
  6. Climate Data         — monthly climate summary (if available)

Dependencies
------------
    pip install openpyxl

Usage
-----
    from ai_engine.export.excel_exporter import ExcelExporter
    from ai_engine.export.pvsyst_exporter import make_synthetic_project

    project = make_synthetic_project()
    exp     = ExcelExporter(project)
    path    = exp.export_workbook('/tmp/shamsi_project.xlsx')
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.styles import (
        Alignment, Border, Fill, Font, GradientFill, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# Colour constants (hex, no leading #)
# ─────────────────────────────────────────────────────────────────────────────

C_BLUE_DARK  = '1e3a8a'
C_BLUE_MID   = '2563eb'
C_BLUE_LIGHT = 'dbeafe'
C_ORANGE     = 'f97316'
C_GREEN      = '16a34a'
C_RED        = 'dc2626'
C_GREY       = 'f3f4f6'
C_WHITE      = 'ffffff'
C_BLACK      = '111827'
C_GREY_TXT   = '6b7280'


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=hex_colour)


def _font(bold=False, size=11, colour=C_BLACK, italic=False) -> Font:
    return Font(bold=bold, size=size, color=colour, italic=italic,
                name='Calibri')


def _align(horiz='left', vert='center', wrap=False) -> Alignment:
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, row: int, cols: int,
                  hex_bg: str = C_BLUE_MID, text_col: str = C_WHITE,
                  font_size: int = 11) -> None:
    """Apply header row styling across *cols* columns starting at column 1."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _fill(hex_bg)
        cell.font      = _font(bold=True, size=font_size, colour=text_col)
        cell.alignment = _align('center')
        cell.border    = _thin_border()


def _data_row_style(ws, row: int, cols: int, even: bool) -> None:
    bg = C_GREY if even else C_WHITE
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _fill(bg)
        cell.font      = _font(size=10)
        cell.alignment = _align()
        cell.border    = _thin_border()


def _kpi_block(ws, start_row: int, label: str, value: str,
               bg: str = C_BLUE_LIGHT, label_col: int = 1) -> None:
    """Write a two-cell KPI block (label | value)."""
    lc = ws.cell(row=start_row, column=label_col, value=label)
    lc.font      = _font(bold=True, size=10, colour=C_BLUE_DARK)
    lc.fill      = _fill(bg)
    lc.alignment = _align()
    lc.border    = _thin_border()

    vc = ws.cell(row=start_row, column=label_col + 1, value=value)
    vc.font      = _font(size=10, colour=C_BLACK)
    vc.fill      = _fill(C_WHITE)
    vc.alignment = _align('right')
    vc.border    = _thin_border()


def _set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _merge_title(ws, row: int, start_col: int, end_col: int,
                 text: str, bg: str = C_BLUE_DARK, font_size: int = 14) -> None:
    ws.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row,   end_column=end_col,
    )
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font      = _font(bold=True, size=font_size, colour=C_WHITE)
    cell.fill      = _fill(bg)
    cell.alignment = _align('center')


def _safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default

def _extract_results(project_data: Dict) -> Dict:
    opt = project_data.get('optimization_results') or {}
    pareto = opt.get('pareto_solutions') or project_data.get('pareto_solutions') or []
    sol = pareto[0] if pareto else {}

    def _pick(*keys, default=0.0):
        for src in [opt, sol, project_data]:
            for k in keys:
                v = src.get(k)
                if v is not None and v != '' and v != '—':
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        return v
        return default

    annual_kwh     = _pick('annual_yield_kwh', 'predicted_annual_kwh', 'annual_kwh', 'annual_production_kwh')
    monthly        = (sol.get('monthly_yield_kwh') or
                      opt.get('monthly_yield_kwh') or
                      opt.get('predicted_monthly') or
                      project_data.get('predicted_monthly') or [])
    if len(monthly) < 12:
        monthly = [annual_kwh / 12.0] * 12 if annual_kwh else [0.0] * 12

    specific_yield = _pick('specific_yield', 'specific_yield_kwh_per_kwp')
    total_cost     = _pick('total_cost_egp', 'cost_egp')
    payback_yrs    = _pick('payback_years', 'payback_period_years')
    annual_savings = _pick('annual_savings_egp', 'annual_saving_egp')
    lifetime_sav   = _pick('lifetime_savings_egp', 'savings_25yr_egp')
    gross_savings  = _pick('gross_savings_25yr', 'total_savings_25yr')
    panel_count    = int(_pick('panel_count', 'num_panels', default=30))
    system_kw      = _pick('system_kw', 'system_kwp', 'capacity_kw')

    # Re-calculate defaults if missing
    p_power = 580.0
    if project_data.get('panel'):
        try:
            p_power = float(getattr(project_data['panel'], 'power_rating_w', 580.0))
        except Exception:
            p_power = 580.0
    if not system_kw and panel_count:
        system_kw = (panel_count * p_power) / 1000.0
    if not specific_yield and system_kw and annual_kwh:
        specific_yield = annual_kwh / system_kw
    if not specific_yield:
        specific_yield = 1750.0
    if not total_cost and system_kw:
        total_cost = system_kw * 18000.0
    cost_per_w     = _pick('cost_per_watt', 'cost_per_wp')
    if not cost_per_w and total_cost and system_kw:
        cost_per_w = total_cost / (system_kw * 1000.0)
    
    if not annual_savings and annual_kwh:
        usage_type = project_data.get('usage_type', 'RESIDENTIAL') or 'RESIDENTIAL'
        tariff = 1.35 if usage_type == 'RESIDENTIAL' else 1.75
        annual_savings = annual_kwh * tariff
        
    if not payback_yrs and total_cost and annual_savings:
        payback_yrs = total_cost / annual_savings
        
    if not lifetime_sav and annual_savings:
        # Net savings over 25 years: cumulative savings - total cost
        degradation = 0.005
        escalation = 0.05
        cum_savings = 0.0
        tariff_price = 1.35 if project_data.get('usage_type', 'RESIDENTIAL') == 'RESIDENTIAL' else 1.75
        for yr in range(1, 26):
            prod_t = annual_kwh * ((1.0 - degradation) ** yr)
            tariff_t = tariff_price * ((1.0 + escalation) ** yr)
            cum_savings += prod_t * tariff_t
        lifetime_sav = cum_savings - total_cost
        gross_savings = cum_savings

    perf_ratio = _pick('performance_ratio', 'pr', default=0.80)
    dust_loss = _pick('dust_loss_pct', default=5.0)

    return {
        'annual_yield_kwh'  : annual_kwh,
        'monthly_yield_kwh' : monthly,
        'specific_yield'    : specific_yield,
        'total_cost_egp'    : total_cost,
        'payback_years'     : payback_yrs,
        'annual_savings_egp': annual_savings,
        'lifetime_savings_egp': lifetime_sav,
        'gross_savings_25yr': gross_savings,
        'panel_count'       : panel_count,
        'system_kw'         : system_kw,
        'performance_ratio' : perf_ratio,
        'cost_per_watt'     : cost_per_w,
        'dust_loss_pct'     : dust_loss,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Main exporter
# ─────────────────────────────────────────────────────────────────────────────

class ExcelExporter:
    """
    Export a Shamsi Smart project to a formatted multi-sheet Excel workbook.

    Parameters
    ----------
    project_data : dict
        Same schema as PVsystExporter.
    """

    MONTHS = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    MONTHS_SHORT = ['Jan','Feb','Mar','Apr','May','Jun',
                    'Jul','Aug','Sep','Oct','Nov','Dec']

    def __init__(self, project_data: Dict):
        self.project  = project_data
        self.location = project_data['location']
        self.panel    = project_data['panel']
        self.inverter = project_data['inverter']
        self.config   = project_data['system_config']
        self.results  = _extract_results(project_data)

    def _validate_data_quality(self) -> None:
        """Verify data quality before exporting to Excel, raising ValueError on failure."""
        res = self.results
        cfg = self.config
        p = self.panel
        inv = self.inverter
        loc = self.location

        # 1. Null or empty required fields
        if not loc or not getattr(loc, 'name', None):
            raise ValueError("Data Validation Error: Location name is missing.")
        if not p or not getattr(p, 'model', None):
            raise ValueError("Data Validation Error: Solar panel specification is missing.")
        if not inv or not getattr(inv, 'model', None):
            raise ValueError("Data Validation Error: Inverter specification is missing.")

        # 2. Key metrics must be positive and non-zero
        if res.get('annual_yield_kwh', 0.0) <= 0.0:
            raise ValueError("Data Validation Error: Annual production must be greater than zero.")
        if res.get('total_cost_egp', 0.0) <= 0.0:
            raise ValueError("Data Validation Error: Total cost of investment must be greater than zero.")
        if res.get('system_kw', 0.0) <= 0.0:
            raise ValueError("Data Validation Error: System capacity must be greater than zero.")

        # 3. Monthly sum equals annual yield within 1 kWh tolerance
        monthly = res.get('monthly_yield_kwh', [])
        if not monthly or len(monthly) < 12:
            raise ValueError("Data Validation Error: Monthly yield profile must contain exactly 12 months.")
        if any(m <= 0 for m in monthly):
            raise ValueError("Data Validation Error: Monthly production values must be strictly positive.")
        
        m_sum = sum(monthly)
        ann_yield = res.get('annual_yield_kwh')
        if abs(m_sum - ann_yield) > 1.0:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(
                f"Excel Export: Sum of monthly production ({m_sum:.2f} kWh) "
                f"does not match the reported annual yield ({ann_yield:.2f} kWh). "
                "Scaling monthly production values to match annual yield."
            )
            scale = ann_yield / m_sum if m_sum > 0 else 1.0
            res['monthly_yield_kwh'] = [m * scale for m in monthly]

        # 4. String configuration checks
        strings = cfg.get('strings')
        pps = cfg.get('panels_per_string')
        total_panels = cfg.get('panel_count')
        if strings and pps and total_panels:
            if strings * pps != total_panels:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(
                    f"Excel Export: Engineering mismatch! String count ({strings}) "
                    f"multiplied by modules per string ({pps}) equals {strings * pps}, "
                    f"which does not match the total panel count ({total_panels})."
                )

        # 5. Financial calculations consistency
        payback = res.get('payback_years', 0.0)
        cost = res.get('total_cost_egp', 0.0)
        savings = res.get('annual_savings_egp', 0.0)
        if cost > 0 and savings > 0:
            expected_payback = cost / savings
            if abs(payback - expected_payback) > 0.5:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(
                    f"Excel Export: Payback period ({payback:.2f} yrs) "
                    f"is inconsistent with investment ({cost:.2f} EGP) and savings ({savings:.2f} EGP/yr)."
                )

    # ── Public API ────────────────────────────────────────────────────────────

    def export_workbook(self, output_file: str) -> str:
        """
        Write formatted Excel workbook to *output_file*.

        Returns
        -------
        str   Absolute path of the written file.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                'openpyxl is required for Excel export. '
                'Install with: pip install openpyxl'
            )

        self._validate_data_quality()

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        self._sheet_summary(wb)
        self._sheet_system_design(wb)
        self._sheet_monthly_production(wb)
        self._sheet_financial_analysis(wb)
        self._sheet_equipment_specs(wb)
        self._sheet_climate_summary(wb)
        self._sheet_validation_warnings(wb)

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if val_str.startswith('='):
                        val_str = "Formula_Result_Length"
                    max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 50)

        # Workbook properties
        wb.properties.title   = f"Shamsi Smart — {self.location.name}"
        wb.properties.creator = "Shamsi Smart AI"
        wb.properties.subject = "Solar PV Design Export"

        wb.save(output_file)
        return os.path.abspath(output_file)

    def export_csv(self, output_file: str) -> str:
        """
        Export a simple CSV with monthly production data.

        Returns
        -------
        str  Absolute path of the written file.
        """
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        monthly = self.results.get('monthly_yield_kwh', [0]*12)
        if len(monthly) < 12:
            monthly = (monthly + [0]*12)[:12]

        lines = ['Month,Production_kWh,Percentage']
        total = sum(monthly)
        for i, (m, v) in enumerate(zip(self.MONTHS, monthly)):
            pct = f"{v/total*100:.1f}" if total > 0 else '0.0'
            lines.append(f"{m},{v:.1f},{pct}")
        lines.append(f"Annual,{total:.1f},100.0")

        with open(output_file, 'w', encoding='utf-8') as fh:
            fh.write('\n'.join(lines))
        return os.path.abspath(output_file)

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────

    def _sheet_summary(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('Summary', 0)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        _set_col_widths(ws, {1: 30, 2: 24, 3: 4, 4: 30, 5: 24})

        res = self.results
        cfg = self.config
        p   = self.panel
        loc = self.location
        sys_kwp = self.project.get('system_kw') or (cfg['panel_count'] * p.power_rating_w / 1000)

        # Title block
        _merge_title(ws, 1, 1, 5, 'SHAMSI SMART AI — SOLAR DESIGN SUMMARY')
        _merge_title(ws, 2, 1, 5,
                     f"{loc.name}  |  {sys_kwp:.2f} kWp  |  "
                     f"Generated {datetime.now().strftime('%d %B %Y')}",
                     bg=C_BLUE_LIGHT, font_size=11)
        ws.cell(row=2, column=1).font = _font(size=11, colour=C_BLUE_DARK, italic=True)

        row = 4
        # Left column KPIs
        left_kpis = [
            ('Location',              loc.name),
            ('Latitude / Longitude',  f"{loc.latitude:.3f}°N, {loc.longitude:.3f}°E"),
            ('System Capacity',       sys_kwp),
            ('Number of Panels',      "='System Design'!B6 * 'System Design'!B7"),
            ('Panel Model',           f"{p.manufacturer} {p.model}"),
            ('Inverter Model',        f"{self.inverter.manufacturer} {self.inverter.model}"),
            ('Tilt / Azimuth',        f"{cfg['tilt_angle']}° / {cfg.get('azimuth', 180)}°"),
        ]
        for label, value in left_kpis:
            _kpi_block(ws, row, label, value, label_col=1)
            row += 1

        row = 4
        # Right column KPIs (Formulas linking to other sheets)
        right_kpis = [
            ('Annual Production',     "='Monthly Production'!B15"),
            ('Specific Yield',        "=E4/B6"),
            ('Total Investment',      res.get('total_cost_egp', 0)),
            ('Cost per Watt',         "=E6/(B6*1000)"),
            ('Simple Payback',        "=E6/'Financial Analysis'!D5"),
            ('25-Year Net Savings',   "='Financial Analysis'!F29"),
            ('CO₂ Avoided / Year',    "=E4*0.48/1000"),
        ]
        for label, value in right_kpis:
            _kpi_block(ws, row, label, value, label_col=4)
            row += 1

        # Apply correct number formatting to formula cells
        ws.cell(row=6, column=2).number_format = '0.000' # System Capacity
        ws.cell(row=7, column=2).number_format = '#,##0' # Number of Panels
        
        ws.cell(row=4, column=5).number_format = '#,##0' # Annual Production
        ws.cell(row=5, column=5).number_format = '#,##0' # Specific Yield
        ws.cell(row=6, column=5).number_format = '#,##0' # Total Investment
        ws.cell(row=7, column=5).number_format = '#,##0.00' # Cost per Watt
        ws.cell(row=8, column=5).number_format = '0.0' # Simple Payback
        ws.cell(row=9, column=5).number_format = '#,##0' # 25-Year Net Savings
        ws.cell(row=10, column=5).number_format = '0.0' # CO₂ Avoided

        # Footer
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1,
                       value='Generated by Shamsi Smart AI — NSGA-II Multi-Objective Optimisation '
                             'with CNN-LSTM Yield Prediction')
        cell.font      = _font(size=9, colour=C_GREY_TXT, italic=True)
        cell.alignment = _align('center')

    def _sheet_system_design(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('System Design', 1)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        _set_col_widths(ws, {1: 32, 2: 28})

        _merge_title(ws, 1, 1, 2, 'System Design Parameters')

        sys_kwp = self.project.get('system_kw') or (self.config['panel_count'] * self.panel.power_rating_w / 1000)
        system_type = self.project.get('system_type_str', 'Grid-Tied Solar System (On-Grid, No Battery)')
        dust_loss = self.project.get('dust_loss_pct', 5.0)

        rows = [
            ('SYSTEM CONFIGURATION', None),
            ('System Type',         system_type),
            ('Peak DC Capacity',    "='Summary'!B6"),
            ('Number of Strings',   self.config.get('strings', 3)),
            ('Modules / String',    self.config.get('panels_per_string', 10)),
            ('Tilt Angle',          self.config.get('tilt_angle', 20.0)),
            ('Azimuth',             self.config.get('azimuth', 180.0)),
            ('Inverter Count',      self.config.get('inverter_count', 1)),
            (None, None),
            ('ROOF LAYOUT DESIGN', None),
            ('Required Roof Area',  "='Summary'!B7 * 'Equipment Specs'!B19 * 1.35"),
            (None, None),
            ('ELECTRICAL CABLE SUMMARY', None),
            ('DC Solar Cable (PV1-F)', self.project.get('cable_summary', {}).get('dc_cable', '—')),
            ('DC Design Voltage Drop', self.project.get('cable_summary', {}).get('dc_voltage_drop', '—')),
            ('AC Main Cable',       self.project.get('cable_summary', {}).get('ac_cable', '—')),
            ('AC Design Voltage Drop', self.project.get('cable_summary', {}).get('ac_voltage_drop', '—')),
            (None, None),
            ('ELECTRICAL PROTECTION DEVICES', None),
            ('DC PV String Fuses',  self.project.get('protections', {}).get('dc_fuse', '—')),
            ('DC Circuit Breaker',  self.project.get('protections', {}).get('dc_breaker', '—')),
            ('DC Surge Protection (SPD)', self.project.get('protections', {}).get('dc_spd', '—')),
            ('AC Circuit Breaker',  self.project.get('protections', {}).get('ac_breaker', '—')),
            ('AC Surge Protection (SPD)', self.project.get('protections', {}).get('ac_spd', '—')),
            (None, None),
            ('LOSS ASSUMPTIONS', None),
            ('Dust / Soiling',      dust_loss / 100.0),
            ('DC Wiring',           0.02),
            ('Mismatch',            0.02),
            ('Inverter',            0.016),
            ('Shading',             float(self.project.get('shading_loss_pct', 3.0)) / 100.0),
            ('Light-Induced Degradation', 0.015),
            ('Annual Degradation',  0.005),
        ]

        for r, (label, value) in enumerate(rows, start=3):
            if label is None:
                continue
            if value is None:
                # Section header
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                cell = ws.cell(row=r, column=1, value=label)
                cell.font = _font(bold=True, size=10, colour=C_WHITE)
                cell.fill = _fill(C_BLUE_DARK)
                cell.alignment = _align()
            else:
                lc = ws.cell(row=r, column=1, value=label)
                vc = ws.cell(row=r, column=2, value=value)
                even = (r % 2 == 0)
                for c in [lc, vc]:
                    c.fill      = _fill(C_GREY if even else C_WHITE)
                    c.font      = _font(size=10)
                    c.alignment = _align()
                    c.border    = _thin_border()
                
                # Apply custom number formatting based on row indices
                if r == 5:
                    vc.number_format = '0.000" kWp"'
                    vc.alignment = _align('right')
                elif r in (6, 7, 10):
                    vc.number_format = '#,##0'
                    vc.alignment = _align('right')
                elif r == 8:
                    vc.number_format = '0.0"°"'
                    vc.alignment = _align('right')
                elif r == 9:
                    vc.number_format = '0.0"° (South=180°)"'
                    vc.alignment = _align('right')
                elif r == 13:
                    vc.number_format = '0.0" m²"'
                    vc.alignment = _align('right')
                elif r in (27, 28, 29, 30, 31, 32):
                    vc.number_format = '0.0%'
                    vc.alignment = _align('right')
                elif r == 33:
                    vc.number_format = '0.0% "/year"'
                    vc.alignment = _align('right')

    # ── Sheet 3: Monthly Production ───────────────────────────────────────────

    def _sheet_monthly_production(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('Monthly Production', 2)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        _set_col_widths(ws, {1: 16, 2: 20, 3: 16, 4: 16})

        monthly = self.results.get('monthly_yield_kwh', [0]*12)
        if len(monthly) < 12:
            monthly = (monthly + [0]*12)[:12]

        _merge_title(ws, 1, 1, 4, 'Monthly Energy Production Forecast')

        # Header row
        headers = ['Month', 'Production (kWh)', '% of Annual', 'Cumulative (kWh)']
        for c, h in enumerate(headers, 1):
            ws.cell(row=2, column=c, value=h)
        _header_style(ws, 2, 4)

        for i, m in enumerate(self.MONTHS):
            row = i + 3
            
            # Formulas
            formula_pct = f"=B{row}/$B$15"
            if i == 0:
                formula_cum = f"=B{row}"
            else:
                formula_cum = f"=D{row-1}+B{row}"
                
            ws.cell(row=row, column=1, value=m)
            ws.cell(row=row, column=2, value=float(monthly[i]))
            ws.cell(row=row, column=3, value=formula_pct)
            ws.cell(row=row, column=4, value=formula_cum)
            
            _data_row_style(ws, row, 4, even=(i % 2 == 0))
            
            # Formatting
            ws.cell(row=row, column=2).number_format = '#,##0.0'
            ws.cell(row=row, column=2).alignment = _align('right')
            ws.cell(row=row, column=3).number_format = '0.0%'
            ws.cell(row=row, column=3).alignment = _align('right')
            ws.cell(row=row, column=4).number_format = '#,##0.0'
            ws.cell(row=row, column=4).alignment = _align('right')

        # Annual total row
        tot_row = 15
        ws.cell(row=tot_row, column=1, value='Annual Total').font = _font(bold=True, size=10)
        ws.cell(row=tot_row, column=2, value="=SUM(B3:B14)").font = _font(bold=True)
        ws.cell(row=tot_row, column=3, value="=SUM(C3:C14)").font = _font(bold=True)
        ws.cell(row=tot_row, column=4, value="=D14").font = _font(bold=True)
        
        for c in range(1, 5):
            ws.cell(row=tot_row, column=c).fill   = _fill(C_BLUE_LIGHT)
            ws.cell(row=tot_row, column=c).border = _thin_border()
            if c >= 2:
                ws.cell(row=tot_row, column=c).alignment = _align('right')
                if c == 3:
                    ws.cell(row=tot_row, column=c).number_format = '0.0%'
                else:
                    ws.cell(row=tot_row, column=c).number_format = '#,##0.0'

        # Bar chart
        data_ref   = Reference(ws, min_col=2, min_row=2, max_row=14)
        cats_ref   = Reference(ws, min_col=1, min_row=3, max_row=14)
        chart      = BarChart()
        chart.type = 'col'
        chart.title            = 'Monthly Energy Production (kWh)'
        chart.y_axis.title     = 'Energy (kWh)'
        chart.x_axis.title     = 'Month'
        chart.style            = 10
        chart.grouping         = 'clustered'
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width  = 22
        chart.height = 12
        chart.series[0].graphicalProperties.solidFill = C_BLUE_MID
        ws.add_chart(chart, 'F2')

    # ── Sheet 4: Financial Analysis ───────────────────────────────────────────

    def _sheet_financial_analysis(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('Financial Analysis', 3)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'
        _set_col_widths(ws, {1: 10, 2: 22, 3: 20, 4: 20, 5: 20, 6: 20})

        _merge_title(ws, 1, 1, 6, '25-Year Financial Cashflow Model')

        # Parameters
        annual_kwh  = self.results.get('annual_yield_kwh', 0) or 0
        total_cost  = self.results.get('total_cost_egp', 0) or 0
        payback_yrs = self.results.get('payback_years', 0) or 0
        
        usage_type = self.project.get('usage_type', 'RESIDENTIAL') or 'RESIDENTIAL'
        elec_price  = 1.35 if usage_type == 'RESIDENTIAL' else 1.75
        degradation = 0.005   # 0.5%/yr
        escalation  = 0.05    # 5%/yr

        # Parameters Block in Excel Columns H & I
        ws.cell(row=3, column=8, value='Design Parameter').font = _font(bold=True, colour=C_BLUE_DARK)
        ws.cell(row=3, column=9, value='Value').font = _font(bold=True, colour=C_BLUE_DARK)
        
        ws.cell(row=4, column=8, value='Annual Degradation')
        ws.cell(row=4, column=9, value=degradation)
        ws.cell(row=4, column=9).number_format = '0.0%'
        
        ws.cell(row=5, column=8, value='Initial Tariff (EGP/kWh)')
        ws.cell(row=5, column=9, value=elec_price)
        ws.cell(row=5, column=9).number_format = '0.00'
        
        ws.cell(row=6, column=8, value='Tariff Escalation Rate')
        ws.cell(row=6, column=9, value=escalation)
        ws.cell(row=6, column=9).number_format = '0.0%'
        
        for r_idx in range(3, 7):
            ws.cell(row=r_idx, column=8).border = _thin_border()
            ws.cell(row=r_idx, column=9).border = _thin_border()
            if r_idx > 3:
                ws.cell(row=r_idx, column=8).fill = _fill(C_GREY)
                ws.cell(row=r_idx, column=9).fill = _fill(C_WHITE)

        # Column headers
        headers = ['Year','Production (kWh)','Price (EGP/kWh)',
                   'Annual Saving (EGP)','Net Cashflow (EGP)','Cumulative (EGP)']
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        _header_style(ws, 3, 6)

        ws.cell(row=4, column=1, value=0)
        ws.cell(row=4, column=5, value="=-Summary!E6")
        ws.cell(row=4, column=6, value="=E4")
        _data_row_style(ws, 4, 6, even=False)
        for c in [5, 6]:
            ws.cell(row=4, column=c).number_format = '#,##0'
            ws.cell(row=4, column=c).alignment = _align('right')

        for yr in range(1, 26):
            row = yr + 4
            ws.cell(row=row, column=1, value=yr)
            ws.cell(row=row, column=2, value=f"='Monthly Production'!$B$15 * (1 - $I$4)^A{row}")
            ws.cell(row=row, column=3, value=f"=$I$5 * (1 + $I$6)^A{row}")
            ws.cell(row=row, column=4, value=f"=B{row} * C{row}")
            ws.cell(row=row, column=5, value=f"=D{row}")
            ws.cell(row=row, column=6, value=f"=F{row-1} + E{row}")
            
            _data_row_style(ws, row, 6, even=(yr % 2 == 0))
            
            # Formatting
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=2).alignment = _align('right')
            ws.cell(row=row, column=3).number_format = '0.000'
            ws.cell(row=row, column=3).alignment = _align('right')
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=4).alignment = _align('right')
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=5).alignment = _align('right')
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=6).alignment = _align('right')

        # Conditional formatting for cumulative column (F4:F29)
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        red_font = Font(color='991B1B', bold=True)
        green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        green_font = Font(color='065F46', bold=True)
        
        ws.conditional_formatting.add('F4:F29', CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))
        ws.conditional_formatting.add('F4:F29', CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=green_fill, font=green_font))

        # Summary KPIs below table
        kpi_row = 31
        ws.merge_cells(start_row=kpi_row, start_column=1, end_row=kpi_row, end_column=6)
        ws.cell(row=kpi_row, column=1, value='Financial Summary').font = _font(bold=True, size=11, colour=C_BLUE_DARK)

        kpis = [
            ('Total Investment',        "=Summary!E6"),
            ('Simple Payback',          "=B32/D5"),
            ('25-Year Gross Savings',   "=SUM(D5:D29)"),
            ('25-Year Net Savings',     "=F29"),
            ('CO₂ Offset (25 yr)',      "=SUM(B5:B29)*0.48/1000"),
        ]
        for i, (label, value) in enumerate(kpis):
            r = kpi_row + 1 + i
            _kpi_block(ws, r, label, value)
            
        ws.cell(row=32, column=2).number_format = '#,##0'
        ws.cell(row=33, column=2).number_format = '0.0'
        ws.cell(row=34, column=2).number_format = '#,##0'
        ws.cell(row=35, column=2).number_format = '#,##0'
        ws.cell(row=36, column=2).number_format = '#,##0.0'

        # Line chart: cumulative cashflow
        cum_ref  = Reference(ws, min_col=6, min_row=3, max_row=29)
        yr_ref   = Reference(ws, min_col=1, min_row=5, max_row=29)
        lc       = LineChart()
        lc.title = 'Cumulative Net Cashflow (EGP)'
        lc.y_axis.title = 'EGP'
        lc.x_axis.title = 'Year'
        lc.style   = 10
        lc.add_data(cum_ref, titles_from_data=True)
        lc.set_categories(yr_ref)
        lc.width   = 22
        lc.height  = 12
        lc.series[0].graphicalProperties.line.solidFill = C_BLUE_MID
        lc.series[0].graphicalProperties.line.width = 20000   # 2pt
        ws.add_chart(lc, 'H3')

    def _sheet_equipment_specs(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('Equipment Specs', 4)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'
        _set_col_widths(ws, {1: 30, 2: 28, 3: 4, 4: 30, 5: 28})

        _merge_title(ws, 1, 1, 5, 'Equipment Specifications')

        p   = self.panel
        inv = self.inverter

        panel_rows = [
            ('SOLAR PANEL', None, None),
            ('Manufacturer',            p.manufacturer, None),
            ('Model',                   p.model, None),
            ('Technology',              p.technology, None),
            ('Rated Power (Pnom)',      _safe_float(p.power_rating_w), '0" W"'),
            ('Efficiency',              _safe_float(p.efficiency_percent) / 100.0, '0.0%'),
            ('Vmpp',                    _safe_float(p.vmp_v), '0.00" V"'),
            ('Impp',                    _safe_float(p.imp_a), '0.00" A"'),
            ('Voc',                     _safe_float(p.voc_v), '0.00" V"'),
            ('Isc',                     _safe_float(p.isc_a), '0.00" A"'),
            ('Temp. Coeff. Pmax',       _safe_float(p.temp_coeff_pmax_percent), '0.000"% / °C"'),
            ('Temp. Coeff. Voc',        _safe_float(p.temp_coeff_voc_percent), '0.000"% / °C"'),
            ('Temp. Coeff. Isc',        _safe_float(p.temp_coeff_isc_percent), '0.000"% / °C"'),
            ('NOCT',                    _safe_float(getattr(p, 'noct_celsius', 45)), '0" °C"'),
            ('Dimensions (L×W×H)',      f"{_safe_float(p.length_mm):.0f} × {_safe_float(p.width_mm):.0f} × {_safe_float(p.thickness_mm):.0f} mm", None),
            ('Weight',                  _safe_float(p.weight_kg), '0.0" kg"'),
            ('Cell Area',               _safe_float(p.area_m2), '0.0000" m²"'),
        ]

        inv_rows = [
            ('INVERTER', None, None),
            ('Manufacturer',            inv.manufacturer, None),
            ('Model',                   inv.model, None),
            ('Type',                    inv.inverter_type, None),
            ('Rated AC Power',          _safe_float(inv.power_rating_w), '#,##0" W"'),
            ('Max AC Power',            _safe_float(inv.max_ac_power_w), '#,##0" W"'),
            ('AC Voltage',              _safe_float(inv.output_voltage_v), '0" V"'),
            ('Max Efficiency',          _safe_float(inv.max_efficiency_percent) / 100.0, '0.0%'),
            ('Euro Efficiency',         _safe_float(inv.euro_efficiency_percent) / 100.0, '0.0%'),
            ('Max DC Voltage',          _safe_float(inv.max_dc_voltage_v), '0" V"'),
            ('Min DC Voltage',          _safe_float(inv.min_dc_voltage_v), '0" V"'),
            ('MPPT Range',              f"{_safe_float(inv.mppt_voltage_min_v):.0f} – {_safe_float(inv.mppt_voltage_max_v):.0f} V", None),
            ('Max DC Current',          _safe_float(inv.max_dc_current_a), '0.0" A"'),
            ('Number of MPPTs',         int(_safe_float(inv.number_of_mppts)), '0'),
            ('Number of DC Inputs',     int(_safe_float(inv.number_of_inputs)), '0'),
            ('Weight',                  _safe_float(inv.weight_kg), '0.0" kg"'),
            ('Dimensions',              str(getattr(inv, 'dimensions_mm', 'N/A')), None),
        ]

        def _write_col(rows, col_offset):
            for r_idx, (label, value, fmt_str) in enumerate(rows, start=3):
                row = r_idx
                if value is None:
                    ws.merge_cells(start_row=row, start_column=col_offset,
                                   end_row=row, end_column=col_offset + 1)
                    cell = ws.cell(row=row, column=col_offset, value=label)
                    cell.fill      = _fill(C_BLUE_DARK)
                    cell.font      = _font(bold=True, colour=C_WHITE, size=10)
                    cell.alignment = _align()
                else:
                    lc = ws.cell(row=row, column=col_offset,   value=label)
                    vc = ws.cell(row=row, column=col_offset+1, value=value)
                    even = (r_idx % 2 == 0)
                    for c in [lc, vc]:
                        c.fill      = _fill(C_GREY if even else C_WHITE)
                        c.font      = _font(size=10)
                        c.border    = _thin_border()
                        c.alignment = _align()
                    if fmt_str:
                        vc.number_format = fmt_str
                        vc.alignment = _align('right')

        _write_col(panel_rows, col_offset=1)
        # Spacer column 3 intentionally blank
        _write_col(inv_rows,   col_offset=4)

    # ── Sheet 6: Climate Summary ──────────────────────────────────────────────

    def _sheet_climate_summary(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('Climate Data', 5)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'
        _set_col_widths(ws, {1: 14, 2: 20, 3: 20, 4: 16, 5: 18, 6: 16, 7: 16})

        _merge_title(ws, 1, 1, 7,
                     f'Monthly Climate Summary — {self.location.name}')

        days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        m_ghi = self.project.get('monthly_ghi') or [150.0]*12
        m_poa = self.project.get('monthly_poa') or [160.0]*12
        m_temp = self.project.get('monthly_temp') or [25.0]*12

        headers = ['Month', 'Avg GHI (kWh/m²/day)', 'Avg POA (kWh/m²/day)', 'Avg Temp (°C)',
                   'Monthly GHI (kWh/m²)', 'Radiation Index', 'Category']
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        _header_style(ws, 3, 7)

        for i, m_name in enumerate(self.MONTHS):
            row     = i + 4
            days    = days_in_month[i]
            avg_ghi = m_ghi[i]
            avg_poa = m_poa[i]
            avg_t   = m_temp[i]
            mon_ghi = m_ghi[i] * days
            ri      = avg_ghi / 6.5

            if ri >= 0.80:
                cat = 'Excellent'
                cat_col = C_GREEN
            elif ri >= 0.60:
                cat = 'Good'
                cat_col = '16a34a'
            elif ri >= 0.40:
                cat = 'Moderate'
                cat_col = 'ca8a04'
            else:
                cat = 'Low'
                cat_col = C_RED

            data = [m_name, round(avg_ghi, 2), round(avg_poa, 2), round(avg_t, 1),
                    round(mon_ghi, 1), round(ri, 3), cat]
            for c, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=c, value=val)
                _data_row_style(ws, row, 7, even=(i % 2 == 0))
                if c >= 2:
                    cell.alignment = _align('right')
                if c == 7:
                    cell.font = _font(size=10, colour=cat_col, bold=True)
                    cell.alignment = _align('center')

        # GHI vs POA Chart
        data_ref = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=15)
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=15)
        chart = BarChart()
        chart.type = 'col'
        chart.title = 'Monthly Solar Irradiance (GHI vs POA)'
        chart.y_axis.title = 'Irradiance (kWh/m²/day)'
        chart.x_axis.title = 'Month'
        chart.style = 10
        chart.grouping = 'clustered'
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 22
        chart.height = 12
        chart.series[0].graphicalProperties.solidFill = '3b82f6'  # GHI Blue
        chart.series[1].graphicalProperties.solidFill = 'f97316'  # POA Orange
        ws.add_chart(chart, 'I3')

    # ── Sheet 7: Validation & Warnings ────────────────────────────────────────

    def _sheet_validation_warnings(self, wb: 'Workbook') -> None:
        ws = wb.create_sheet('Validation & Warnings', 6)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'
        _set_col_widths(ws, {1: 30, 2: 24, 3: 28, 4: 16, 5: 50})

        _merge_title(ws, 1, 1, 5, 'ENGINEERING VALIDATION & CODE COMPLIANCE')

        p = self.panel
        inv = self.inverter
        cfg = self.config
        res = self.results

        # Calculate values for validations
        p_power = _safe_float(p.power_rating_w)
        panel_count = int(cfg.get('panel_count', 30))
        system_kw = (panel_count * p_power) / 1000.0

        inv_kw = _safe_float(inv.power_rating_w) / 1000.0
        inv_count = int(cfg.get('inverter_count', 1))
        dc_ac_ratio = system_kw / (inv_kw * inv_count) if (inv_kw and inv_count) else 1.0

        panels_per_string = int(cfg.get('panels_per_string', 10))
        panel_voc = _safe_float(p.voc_v, default=50.26)
        temp_coeff_voc = _safe_float(p.temp_coeff_voc_percent, default=-0.27)
        t_min = 5.0
        voc_cold = panel_voc * (1.0 + (t_min - 25.0) * temp_coeff_voc / 100.0)
        string_voc_cold = panels_per_string * voc_cold
        max_dc_v = _safe_float(inv.max_dc_voltage_v, default=1000.0)

        panel_vmp = _safe_float(p.vmp_v, default=41.88)
        temp_coeff_pmax = _safe_float(p.temp_coeff_pmax_percent, default=-0.35)
        t_max = 70.0
        vmp_hot = panel_vmp * (1.0 + (t_max - 25.0) * temp_coeff_pmax / 100.0)
        string_vmp_hot = panels_per_string * vmp_hot
        mppt_min_v = _safe_float(inv.mppt_voltage_min_v, default=200.0)

        inv_type = getattr(inv, 'inverter_type', 'ON_GRID')
        pr = res.get('performance_ratio', 0.80)
        dust_loss = res.get('dust_loss_pct', 5.0)
        shading_loss = float(self.project.get('shading_loss_pct', 3.0))

        # Header row
        headers = ['Compliance Check', 'Operating Value', 'Design Target / Rule', 'Status', 'Recommendation / Details']
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        _header_style(ws, 3, 5)

        # Write rows
        ws.cell(row=4, column=1, value='DC/AC Capacity Ratio')
        ws.cell(row=4, column=2, value=dc_ac_ratio)
        ws.cell(row=4, column=2).number_format = '0.00'
        ws.cell(row=4, column=3, value='1.00 – 1.35')
        ws.cell(row=4, column=4, value='=IF(AND(B4>=1.0, B4<=1.35), "PASS", "WARNING")')
        ws.cell(row=4, column=5, value='=IF(B4<1.0, "Inverter capacity is underutilized. Consider expanding DC module sizing.", IF(B4>1.35, "Inverter capacity is undersized. DC clipping may occur during peak hours.", "DC/AC capacity sizing is optimal."))')

        ws.cell(row=5, column=1, value='Cold Voc Overvoltage safety')
        ws.cell(row=5, column=2, value=string_voc_cold)
        ws.cell(row=5, column=2).number_format = '0.0" V"'
        ws.cell(row=5, column=3, value=f"< {max_dc_v:.0f} V (Max DC Voltage)")
        ws.cell(row=5, column=4, value="=IF(B5<'Equipment Specs'!E13, \"PASS\", \"FAIL\")")
        ws.cell(row=5, column=5, value='=IF(D5="PASS", "String Voc is within safe limits for winter conditions.", "CRITICAL: String open-circuit voltage exceeds inverter maximum rating!")')

        ws.cell(row=6, column=1, value='Hot MPPT Voltage check')
        ws.cell(row=6, column=2, value=string_vmp_hot)
        ws.cell(row=6, column=2).number_format = '0.0" V"'
        ws.cell(row=6, column=3, value=f"> {mppt_min_v:.0f} V (Min MPPT Voltage)")
        ws.cell(row=6, column=4, value="=IF(B6>='Equipment Specs'!E15, \"PASS\", \"WARNING\")")
        ws.cell(row=6, column=5, value='=IF(D6="PASS", "MPPT voltage range is optimal for hot summer conditions.", "Voltage drops below minimum tracker range. Tracker efficiency will degrade.")')

        ws.cell(row=7, column=1, value='Inverter Classification check')
        ws.cell(row=7, column=2, value=inv_type)
        ws.cell(row=7, column=3, value='Matches Battery Setup')
        ws.cell(row=7, column=4, value='=IF(OR(AND(\'System Design\'!B4="Grid-Tied Solar System (On-Grid, No Battery)", B7<>"OFF_GRID"), AND(\'System Design\'!B4="Off-Grid Solar System (with Battery Storage)", B7="OFF_GRID")), "PASS", "WARNING")')
        ws.cell(row=7, column=5, value='=IF(D7="PASS", "Inverter type is compatible with system topology.", "System / Inverter topology mismatch. Verify battery backing configuration.")')

        ws.cell(row=8, column=1, value='Performance Ratio validation')
        ws.cell(row=8, column=2, value=pr)
        ws.cell(row=8, column=2).number_format = '0.0%'
        ws.cell(row=8, column=3, value='75.0% – 86.0%')
        ws.cell(row=8, column=4, value='=IF(AND(B8>=0.75, B8<=0.86), "PASS", "FAIL")')
        ws.cell(row=8, column=5, value='=IF(D8="PASS", "Performance Ratio is within normal range.", "PR is abnormally low or high. Verify shading and temperature coefficients.")')

        for row in range(4, 9):
            _data_row_style(ws, row, 5, even=(row % 2 == 0))
            ws.cell(row=row, column=4).alignment = _align('center')

        # Formatting status results (PASS/WARNING/FAIL)
        green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        green_font = Font(color='065F46', bold=True)
        yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        yellow_font = Font(color='92400E', bold=True)
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        red_font = Font(color='991B1B', bold=True)

        ws.conditional_formatting.add('D4:D8', CellIsRule(operator='equal', formula=['"PASS"'], fill=green_fill, font=green_font))
        ws.conditional_formatting.add('D4:D8', CellIsRule(operator='equal', formula=['"WARNING"'], fill=yellow_fill, font=yellow_font))
        ws.conditional_formatting.add('D4:D8', CellIsRule(operator='equal', formula=['"FAIL"'], fill=red_fill, font=red_font))

        # Overall Status
        ws.cell(row=10, column=1, value='Sizing Review Status').font = _font(bold=True)
        ws.cell(row=10, column=2, value='=IF(COUNTIF(D4:D8, "FAIL") > 0, "CRITICAL VIOLATIONS DETECTED", IF(COUNTIF(D4:D8, "WARNING") > 0, "WARNINGS DETECTED", "DESIGN COMPLIANT"))').font = _font(bold=True)
        ws.cell(row=10, column=2).alignment = _align('center')
        ws.cell(row=10, column=3, value='=IF(COUNTIF(D4:D8, "FAIL") > 0, "Redesign system configurations to prevent hardware damage.", IF(COUNTIF(D4:D8, "WARNING") > 0, "Review warnings to optimize performance.", "System is fully certified for grid connection."))')
        ws.merge_cells(start_row=10, start_column=3, end_row=10, end_column=5)

        for col_idx in range(1, 6):
            cell = ws.cell(row=10, column=col_idx)
            cell.border = _thin_border()
            cell.fill = _fill(C_BLUE_LIGHT)

        ws.conditional_formatting.add('B10', CellIsRule(operator='equal', formula=['"DESIGN COMPLIANT"'], fill=green_fill, font=green_font))
        ws.conditional_formatting.add('B10', CellIsRule(operator='equal', formula=['"WARNINGS DETECTED"'], fill=yellow_fill, font=yellow_font))
        ws.conditional_formatting.add('B10', CellIsRule(operator='equal', formula=['"CRITICAL VIOLATIONS DETECTED"'], fill=red_fill, font=red_font))

        # Loss breakdown summary
        ws.cell(row=12, column=1, value='SYSTEM LOSS BUDGET').font = _font(bold=True, colour=C_WHITE)
        ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=2)
        ws.cell(row=12, column=1).fill = _fill(C_BLUE_DARK)
        ws.cell(row=12, column=2).fill = _fill(C_BLUE_DARK)
        ws.cell(row=12, column=1).border = _thin_border()
        ws.cell(row=12, column=2).border = _thin_border()

        losses = [
            ('Dust & Soiling', dust_loss / 100.0),
            ('DC Wiring (Ohmic)', 0.02),
            ('Mismatch Loss', 0.02),
            ('Inverter Conversion', 0.016),
            ('Shading Loss', shading_loss / 100.0),
        ]
        for idx, (label, val) in enumerate(losses):
            r = 13 + idx
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            _data_row_style(ws, r, 2, even=(idx % 2 == 0))
            ws.cell(row=r, column=2).number_format = '0.0%'
            ws.cell(row=r, column=2).alignment = _align('right')

        # Total estimated loss
        r_tot = 18
        ws.cell(row=r_tot, column=1, value='Total Estimated Losses').font = _font(bold=True)
        ws.cell(row=r_tot, column=2, value='=1-(1-B13)*(1-B14)*(1-B15)*(1-B16)*(1-B17)').font = _font(bold=True)
        ws.cell(row=r_tot, column=1).fill = _fill(C_BLUE_LIGHT)
        ws.cell(row=r_tot, column=2).fill = _fill(C_BLUE_LIGHT)
        ws.cell(row=r_tot, column=1).border = _thin_border()
        ws.cell(row=r_tot, column=2).border = _thin_border()
        ws.cell(row=r_tot, column=2).number_format = '0.0%'
        ws.cell(row=r_tot, column=2).alignment = _align('right')

        # Pie Chart for losses
        pie = PieChart()
        labels_ref = Reference(ws, min_col=1, min_row=13, max_row=17)
        data_ref = Reference(ws, min_col=2, min_row=12, max_row=17)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.title = "System Loss Budget Breakdown"
        pie.width = 16
        pie.height = 10
        ws.add_chart(pie, 'D12')
